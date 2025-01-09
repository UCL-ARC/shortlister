from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List
import csv
import pickle
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border
import pymupdf
import re

from tabulate import tabulate


@dataclass(frozen=True)
class Criterion:
    """A property of Role - contained within the attribute criteria(list of Criterion objects)."""
    name: str
    description: str

    def __repr__(self) -> str:
        return self.name


@dataclass(eq=False)
class Applicant:
    """A property of Shortlist - contained within the attribute applicants(list of Applicant objects)."""
    name: str = field(compare=True)
    cv: Path
    email: str
    phone: str
    postcode: str
    country_region: str
    right_to_work: bool
    visa_requirement: str
    application_text: str
    scores: Dict[Criterion, str]
    notes: str

    def __repr__(self):
        return f"Applicant(name={self.name}, scores={self.scores}, notes={self.notes})"

    def __hash__(self):
        return hash(self.cv)


@dataclass
class Role:
    """A property of Shortlist."""
    job_title: str
    job_id: str
    criteria: List[Criterion]


@dataclass
class Shortlist:
    """Major class object containing all relevant role, applicant, criteria information for shortlisting."""
    role: Role
    applicants: List[Applicant]


# Constant variables

PICKLE_FILE_NAME = "shortlist.pickle"
CRITERIA_FILE_NAME = "criteria.csv"
RANK_AND_SCORE = {
    "Unsatisfactory": 0,
    "Moderate": 10,
    "Satisfactory": 20,
    "Excellent": 40,
}

# Functions

def load_pickle(file_path):
    """Load shortlist from existing pickle file."""
    with open(file_path, "rb") as f:
        shortlist = pickle.load(f)
    return shortlist


def save_shortlist(path, shortlist):
    """Save shortlist as a pickle file within the role directory path."""
    with open(path / PICKLE_FILE_NAME, "wb") as f:
        pickle.dump(shortlist, f)


def load_shortlist(path: Path) -> (Shortlist, str):
    """Import shortlist data from either: 1. Pickle file or 2. Role directory (when there is no existing pickle data)."""
    file = path / PICKLE_FILE_NAME
    if file.exists():
        shortlist = load_pickle(file)
        msg = "SHORTLIST RESTORED FROM FILE"
    else:
        criteria = load_criteria(path / CRITERIA_FILE_NAME)
        role = load_role(path, criteria)
        applicants = load_applicants(path)
        shortlist = Shortlist(role, applicants)
        msg = "SHORTLIST CREATED FROM PACKS"

    return shortlist, msg


def load_role(path, criteria):
    """Generates role object instance."""
    role = Role(str(path), "0001", criteria)
    return role


def load_applicants(path: Path):
    """Generate a list of applicant instances from pdf format CVs."""
    files = path.glob("*.pdf")
    applicants = []
    for file in files:
        applicant = load_applicants_from_pdf(file)
        applicants.append(applicant)

    sort_alpha(applicants)
    return applicants


def load_applicants_from_pdf(file: Path):
    """Create a single Applicant instance from candidate pack within the parsed in PDF file"""
    doc = pymupdf.open(file)

    # takes the first page of the pdf (the candidate pack)
    cover = doc[0]
    rest_of_pages = doc[1:]
    # extract the remaining pdf pages
    remaining_pdf = "\n".join([page.get_text(sort=True) for page in rest_of_pages])
    # extract text in reading order
    text = cover.get_text(sort=True)
    # turns text into a list of string representing each extracted line
    lines = text.splitlines()
    # remove empty line from list
    cleaned = [line.strip() for line in lines if len(line.strip())]

    # sets the value of each field
    info = extract_info_from_text(cleaned)

    applicant = Applicant(
        name=f'{info["First Name"]} {info["Last Name"]}',
        cv=file,
        email=info["Email Address"],
        phone=info["Preferred Phone Number"],
        postcode=info["Postcode"],
        country_region=info["Country & Region"],
        right_to_work=bool(info["Right To Work"]),
        visa_requirement=info["Visa Requirements"],
        application_text=remaining_pdf,
        scores={},
        notes="",
    )

    # if either of the name field can't be extracted, get applicant's name from their cv filename
    if "<missing>" in applicant.name:
        name_parts = file.stem.split("_")
        applicant.name = " ".join(name_parts[0:2])
        print(f"ERROR READING: {file.name}")

    return applicant


def load_criteria(csv_file):
    """Generate criteria(list of criterion instances) from csv file."""
    criteria = []
    with open(csv_file, "r", encoding="UTF-8") as file:
        reader = csv.reader(file)
        next(reader)

        for row in reader:
            criterion = Criterion(name=row[0], description=row[1])
            criteria.append(criterion)
    return criteria


def update_applicant_score(
    applicant: Applicant, criterion: Criterion, score_index: int
):
    """Updates applicant's score field with selected criterion and selected score"""
    applicant.scores[criterion] = list(RANK_AND_SCORE)[score_index]


def update_applicant_notes(applicant: Applicant, new_note: str):
    """Appends new note to applicant's notes section."""

    if len(applicant.notes):
        applicant.notes += "; "

    applicant.notes += new_note


def total_score(scores: Dict[Criterion, str]) -> int:
    """Takes applicant scores dictionary and returns a total score as a single number"""

    values = [RANK_AND_SCORE.get(score) for score in scores.values()]
    return sum(values)


def sort_alpha(applicants: List[Applicant]):
    """Sort by alphabetical order."""
    return applicants.sort(key=lambda applicant: applicant.name)


def sort_ascending_score(applicants: List[Applicant]):
    """Sort by score (lowest to highest)."""
    return applicants.sort(key=lambda applicant: total_score(applicant.scores))


def sort_descending_score(applicants: List[Applicant]):
    """Sort by score(highest to lowest)."""
    return applicants.sort(
        reverse=True, key=lambda applicant: total_score(applicant.scores)
    )


def clear_score(applicant: Applicant, criterion: Criterion):
    """Removes criterion from Applicant's scores dictionary."""
    if criterion in applicant.scores:
        del applicant.scores[criterion]


# text extraction


def extract_info_from_text(lines: List[str]):
    """gets the section containing applicant information from extracted text"""

    # fields labels to get related applicant information
    labels: Dict[str, str|bool] = dict.fromkeys(
        [
            "First Name",
            "Last Name",
            "Email Address",
            "Preferred Phone Number",
            "Postcode",
            "Country & Region",
            "Right To Work",
            "Visa Requirements",
        ],
        "<missing>",
    )

    # removes header/footer and other irrelevant info
    applicant_info = lines[1:-5]
    right_to_work = lines[-5:-1]

    # filter out the field name and retain only the info to applicant
    for label in labels:
        for line in applicant_info:
            if line.startswith(label):
                data = line.removeprefix(
                    label
                )  # removes the field and leaves only the information
                labels[label] = data.strip()  # removes whitespaces
                break
        else:
            continue

    # finds where the question is and checks the next index which contains the answer to the question
    applicant_right_to_work = None
    visa_req_text = None
    if "Do you have the unrestricted right to work in the UK?" in right_to_work:
        i = right_to_work.index("Do you have the unrestricted right to work in the UK?")
        if right_to_work[i + 1] == "No":
            j = right_to_work.index(
                "If no, please give details of your VISA requirements"
            )
            applicant_right_to_work = False
            visa_req_text = right_to_work[j + 1]

        elif right_to_work[i + 1] == "Yes":
            applicant_right_to_work = True
            visa_req_text = None

        labels["Right To Work"] = applicant_right_to_work
        labels["Visa Requirements"] = visa_req_text

    return labels


# creating tabular data

def applicant_table(applicants: List[Applicant], criteria: List[Criterion], table_type="wide") -> (List, List):
    """Generates applicant and score data for summary table"""
    # tab is a list of lists:
    # each list in tab has the format of ["1","name1","score1","score2","score3","score*n"]
    tab = []

    # creates heading
#
    header = ["№", "NAME", "Σ"]
    if table_type == "wide":
        criteria_headings = abbreviate([criterion.name for criterion in criteria])
        header = header + criteria_headings
    else:
        header = header + ["SCORES"]

    i = 0  # sets the applicant number
    for applicant in applicants:
        i += 1
        row = [i]
        if len(applicant.name) > 15:
            row.append(applicant.name[0:15] + "...")
        else:
            row.append(applicant.name)

        row.append(total_score(applicant.scores))

        scores = []
        for criterion in criteria:
            if criterion in applicant.scores:
                scores.append(applicant.scores.get(criterion)[0])
            else:
                scores.append("·")

        if table_type == "wide":
            row += scores
        else:
            row += ["".join(scores)]
        tab.append(row)

    table = tabulate(
        tab,
        headers=header,
        stralign="center",
        colalign=("center", "left"),
    )
    return table

def export_excel(applicants: List[Applicant], criteria: List[Criterion]):
        """Save selected applicant(s) data to Excel spreadsheet"""
        # create an instance of workbook
        wb = Workbook()
        # select the first worksheet as active sheet
        ws = wb.active
        ws.title = "selected_applicants"

        # header
        header = ["№", "NAME", "Σ","Right to Work"]
        criteria_headings = [criterion.name for criterion in criteria]
        header = header + criteria_headings
        ws.append(header)

        i = 0
        # rest of applicant information
        for applicant in applicants:
            i = i+1
            row = [i,applicant.name,total_score(applicant.scores),applicant.right_to_work]

            for criterion in criteria:
                if criterion in applicant.scores:
                    row.append(applicant.scores.get(criterion)[0])
                else:
                    row.append("·")

            ws.append(row)

        # Styling
        # Auto adjust width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length+1.5) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # change header to bold
        for col in range(1,len(header)+1):
            ws[get_column_letter(col)+"1"].font = Font(bold=True)
            ws[get_column_letter(col)+"1"].fill = PatternFill(start_color='B7DEE8', fill_type="solid")

        # add colour for cells depending on the score: U(red),M(yellow),S,E(green)
        for row in ws.iter_rows(2):
            for cell in row:
                if cell.value == "U":
                    cell.fill = PatternFill(start_color='FF3300', fill_type="solid")
                elif cell.value == "M":
                    cell.fill = PatternFill(start_color='FFFF00', fill_type="solid")
                elif cell.value == "S":
                    cell.fill = PatternFill(start_color='C4D79B', fill_type="solid")
                elif cell.value == "E":
                    cell.fill = PatternFill(start_color='92D050', fill_type="solid")
                    

        wb.save("spreadsheet.xlsx")


def abbreviate(list_of_strings: List[str]) -> list[str]:
    """Create abbreviations for all strings in a list."""
    abbreviations = []
    for string in list_of_strings:
        if " " in string:
            separated = string.split(
                " "
            )  # separates individual words in string into a list
            abbrev = "".join(
                word[0].upper() for word in separated
            )  # combine intial of all words and return as uppercase
            abbreviations.append(abbrev)
        else:
            abbreviations.append(string)
    return abbreviations


# filter functions
def name(applicant: Applicant, _name: str):
    """Filter by matching name pattern applicant name.
    Example usage:  name(applicant,"Emma")"""
    return re.search(_name, applicant.name)


def score(applicant: Applicant, _name, _score):
    """Filter by matching applicant score.
    Example usage:  score(applicant,"PhD","Excellent")"""

    # checks that (criterion) name does not match any criterion in applicant scores
    if _score is None:
        return _name.lower() not in [
            getattr(criterion, "name").lower() for criterion in applicant.scores
        ]

    # checks if (criterion) name and score matches the saved applicant scores
    for criterion in applicant.scores:
        if getattr(criterion, "name").lower() == _name.lower():
            return applicant.scores[criterion].lower() == _score.lower()
    # handles no match cases
    return False


def rtw(applicant: Applicant):
    """Filter out applicants without the right to work.
    Example usage:  rtw(applicant)"""
    return applicant.right_to_work


def cv(applicant: Applicant, pattern: str):
    """Filter by matching regex pattern in applicant's CV.
    Example usage:  cv(applicant,"Engineer")"""
    return re.search(pattern, applicant.application_text)


def notes(applicant: Applicant, pattern: str):
    """Filter by matching regex pattern in applicant note.
    Example usage:  notes(applicant,"Engineer")
    """
    return re.search(pattern, applicant.notes)


class InteractiveSorter:
    def __init__(self):
        self.selected = None
        self.sorted = None

    def sort(self, items):
        """Insertion sort, but yields the items to compare. The caller says which is greater."""

        # We sort the items inplace, so don't change the input
        self.sorted = items.copy()

        # Insertion sort from https://www.w3schools.com/dsa/dsa_algo_insertionsort.php
        for i in range(1, len(self.sorted)):
            insert_index = i
            value = self.sorted[i]
            for j in range(i - 1, -1, -1):
                # self.selected is set by the caller for the yielded pair
                self.selected = None
                yield self.sorted[j], value
                if self.selected == self.sorted[j]:
                    self.sorted[j + 1] = self.sorted[j]
                    insert_index = j
                else:
                    break
            self.sorted[insert_index] = value
